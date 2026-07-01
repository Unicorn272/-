import io
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY       = "1B3A6B"
NAVY_MED   = "2E5090"
GRAY_BG    = "F2F2F2"
WHITE      = "FFFFFF"
TEAL_TEXT  = "1E6FA5"
GRAY_TEXT  = "888888"

_thin        = Side(style="thin", color="C8C8C8")
_thin_black  = Side(style="thin", color="000000")
_border      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_border_black = Border(left=_thin_black, right=_thin_black,
                       top=_thin_black, bottom=_thin_black)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, italic=False, color="000000", size=11) -> Font:
    return Font(bold=bold, italic=italic, color=color, size=size)


def _align(wrap=True, h="left", indent=0) -> Alignment:
    return Alignment(wrap_text=wrap, vertical="center", horizontal=h, indent=indent)


def _row_h(*args: tuple) -> int:
    """(텍스트, 열너비) 쌍 → 행 높이 추정."""
    max_lines = 1
    for text, width in args:
        if text and width > 0:
            lines = max(1, (len(str(text)) + width - 1) // width)
            max_lines = max(max_lines, lines)
    return max(20, min(max_lines * 15, 300))


def _is_markdown_table(text: str) -> bool:
    if not text:
        return False
    lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
    return len(lines) >= 2 and lines[0].startswith("|") and lines[1].startswith("|")


def _parse_md_table(text: str) -> list[list[str]]:
    rows = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        if re.fullmatch(r"[\|\s\-:]+", line):
            continue
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if any(cells):
            rows.append(cells)
    if not rows:
        return []
    col_count = max(len(r) for r in rows)
    return [r + [""] * (col_count - len(r)) for r in rows]


def _build_sheet1(wb: Workbook, sections: list[dict]):
    ws = wb.active
    ws.title = "덱 구성"
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 58
    ws.sheet_properties.outlinePr.summaryBelow = False

    ws.row_dimensions[1].height = 22
    for col, text in enumerate(["Section", "장", "제목", "내용/핵심메시지"], start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = _align()
        cell.border = _border

    row = 2
    for sec in sections:
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1,
                       value=f"Section {sec.get('section_no', '')}  {sec.get('section_title', '')}")
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = _align(indent=1)
        cell.border = _border
        row += 1

        for ch in sec.get("chapters", []):
            title = ch.get("title", "")
            desc  = ch.get("description", "")
            ws.row_dimensions[row].height = _row_h((title, 32), (desc, 58))
            ws.row_dimensions[row].outlineLevel = 1
            ws.row_dimensions[row].hidden = True
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = _fill(WHITE)
                ws.cell(row=row, column=c).border = _border

            no_cell = ws.cell(row=row, column=2, value=f"{ch.get('no', '')}장")
            no_cell.alignment = _align(h="center")
            no_cell.font = _font(color="555555", size=10)

            t = ws.cell(row=row, column=3, value=title)
            t.font = _font(bold=True, size=11)
            t.alignment = _align(indent=1)

            d = ws.cell(row=row, column=4, value=desc)
            d.alignment = _align()
            d.font = _font(color="333333", size=10)
            row += 1

        key_msg = sec.get("key_message", "")
        ws.row_dimensions[row].height = _row_h((key_msg, 105))
        ws.row_dimensions[row].outlineLevel = 1
        ws.row_dimensions[row].hidden = True
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1, value=key_msg)
        cell.fill = _fill(GRAY_BG)
        cell.font = _font(italic=True, color="444444", size=10)
        cell.alignment = _align(indent=2)
        cell.border = _border
        row += 2


def _write_data_item(ws, row: int, item: dict) -> int:
    """data 항목 1개 렌더링. 사용한 마지막 row 반환."""
    text_val   = item.get("text", "")
    data_val   = item.get("data", "")
    source_val = item.get("source", "")

    if _is_markdown_table(data_val):
        # 제목 행
        ws.row_dimensions[row].height = _row_h((text_val, 108))
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value=text_val)
        cell.font = _font(color=TEAL_TEXT, size=10)
        cell.alignment = _align(indent=1)
        cell.border = _border
        row += 1

        tbl_rows = _parse_md_table(data_val)
        if tbl_rows:
            col_count = len(tbl_rows[0])
            for j in range(3, col_count):
                ws.column_dimensions[get_column_letter(j + 1)].width = 18

            ws.row_dimensions[row].height = 22
            for j, val in enumerate(tbl_rows[0]):
                cell = ws.cell(row=row, column=j + 1, value=val)
                cell.fill = _fill(GRAY_BG)
                cell.font = _font(bold=True, color="444444", size=10)
                cell.alignment = _align(h="center")
                cell.border = _border_black
            row += 1

            for data_row in tbl_rows[1:]:
                max_cell = max(data_row, key=len, default="")
                ws.row_dimensions[row].height = _row_h((max_cell, 18))
                for j, val in enumerate(data_row):
                    cell = ws.cell(row=row, column=j + 1, value=val)
                    cell.fill = _fill(WHITE)
                    cell.alignment = _align(h="center")
                    cell.font = _font(size=10)
                    cell.border = _border_black
                for j in range(len(data_row), max(3, col_count)):
                    ws.cell(row=row, column=j + 1).fill = _fill(WHITE)
                    ws.cell(row=row, column=j + 1).border = _border_black
                row += 1

        if source_val:
            ws.row_dimensions[row].height = _row_h((source_val, 108))
            ws.merge_cells(f"A{row}:C{row}")
            cell = ws.cell(row=row, column=1, value=f"출처: {source_val}")
            cell.font = _font(italic=True, color=GRAY_TEXT, size=9)
            cell.alignment = _align(indent=1)
            cell.border = _border
            row += 1

        row += 1  # 테이블 아이템 후 여백

    else:
        ws.row_dimensions[row].height = _row_h((text_val, 36), (data_val, 42), (source_val, 30))
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = _fill(WHITE)
            ws.cell(row=row, column=c).border = _border

        t = ws.cell(row=row, column=1, value=text_val)
        t.alignment = _align()
        t.font = _font(size=11)

        d = ws.cell(row=row, column=2, value=data_val)
        d.alignment = _align()
        d.font = _font(size=10)

        s = ws.cell(row=row, column=3, value=source_val)
        s.alignment = _align()
        s.font = _font(italic=True, color=GRAY_TEXT, size=10)
        row += 1

    return row


def _build_sheet2(wb: Workbook, sections: list[dict], data_limitations: list[str]):
    ws = wb.create_sheet(title="리서치 데이터")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 30
    ws.sheet_properties.outlinePr.summaryBelow = False

    ws.row_dimensions[1].height = 22
    for col, text in enumerate(["항목", "데이터/수치 (신고서 원문)", "출처"], start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = _align()
        cell.border = _border

    row = 2
    for sec in sections:
        # Section 헤더 (visible, 그룹 기준 행)
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1,
                       value=f"Section {sec.get('section_no', '')}  {sec.get('section_title', '')}")
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = _align(indent=1)
        cell.border = _border
        row += 1

        for ch in sec.get("chapters", []):
            # 장 헤더 - level 1 그룹 (기본 collapsed)
            ws.row_dimensions[row].height = 22
            ws.row_dimensions[row].outlineLevel = 1
            ws.row_dimensions[row].hidden = True
            ws.merge_cells(f"A{row}:C{row}")
            cell = ws.cell(row=row, column=1,
                           value=f"{ch.get('no', '')}장  {ch.get('title', '')}")
            cell.fill = _fill(NAVY_MED)
            cell.font = _font(bold=True, color=WHITE)
            cell.alignment = _align(indent=1)
            cell.border = _border
            row += 1

            for item in ch.get("data", []):
                data_start = row
                row = _write_data_item(ws, row, item)
                # 데이터 행 - level 2 그룹 (기본 collapsed)
                for r in range(data_start, row):
                    ws.row_dimensions[r].outlineLevel = 2
                    ws.row_dimensions[r].hidden = True

        row += 1  # 섹션 간 여백

    if data_limitations:
        row += 1
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value="⚠ 데이터 한계 및 유의사항")
        cell.fill = _fill(GRAY_BG)
        cell.font = _font(bold=True, size=11)
        cell.alignment = _align()
        cell.border = _border
        row += 1
        for limit in data_limitations:
            ws.row_dimensions[row].height = _row_h((limit, 108))
            ws.merge_cells(f"A{row}:C{row}")
            cell = ws.cell(row=row, column=1, value=f"• {limit}")
            cell.alignment = _align(indent=1)
            cell.font = _font(color="555555", size=10)
            cell.border = _border
            row += 1


def generate_excel(analysis: dict, query: str, analysis_date: str) -> bytes:
    sections = analysis.get("sections", [])
    wb = Workbook()
    _build_sheet1(wb, sections)
    _build_sheet2(wb, sections, analysis.get("data_limitations", []))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
